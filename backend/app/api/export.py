from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from app.database import get_db
from app.services.summary import calc_rollup, calc_dept_summary
from app.utils.auth import get_current_user, require_role

router = APIRouter(prefix="/api/export", tags=["Excel导出"])


def build_rollup_excel(db: Session, dept_code: str, year: int, period_type: str):
    data, subject_map, sorted_codes = calc_rollup(db, dept_code, year, period_type)

    wb = Workbook()
    ws = wb.active
    ws.title = "推移表"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    headers = ["科目代码", "科目名称", "类别"]
    for m in range(1, 13):
        headers.append(f"{m}月")
    headers.append("年度累计")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    categories_order = ["收入", "变动费用", "边界利益", "固定费用", "附加价值", "时间", "重要指标", "人工费"]
    row = 2
    for cat in categories_order:
        cat_codes = [c for c in sorted_codes if subject_map.get(c, {}).get("category") == cat]
        if not cat_codes:
            continue
        for code in cat_codes:
            info = subject_map.get(code, {})
            ws.cell(row=row, column=1, value=code).border = thin_border
            ws.cell(row=row, column=2, value=info.get("name", "")).border = thin_border
            ws.cell(row=row, column=3, value=cat).border = thin_border
            for m_idx in range(12):
                month_data = data[m_idx] if m_idx < len(data) else {}
                val = month_data.get("subjects", {}).get(code)
                cell = ws.cell(row=row, column=4 + m_idx, value=val)
                cell.border = thin_border
                cell.alignment = center_align
                if val is not None:
                    cell.number_format = '#,##0.00'
            cumulative_val = data[-1].get("cumulative", {}).get(code) if data else None
            cum_cell = ws.cell(row=row, column=16, value=cumulative_val)
            cum_cell.border = thin_border
            cum_cell.alignment = center_align
            if cumulative_val is not None:
                cum_cell.number_format = '#,##0.00'
            row += 1

    for col_idx in range(1, 17):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.get("/rollup-report")
def export_rollup(
    dept_code: str = Query(...),
    year: int = Query(...),
    period_type: str = Query("actual"),
    token: str = Query(None),
    db: Session = Depends(get_db),
):
    from app.utils.auth import decode_access_token
    from app.models.models import User

    if token:
        payload = decode_access_token(token)
        if not payload:
            from fastapi import HTTPException
            raise HTTPException(status_code=401, detail="Token无效")
        user_id = payload.get("sub")
        current_user = db.query(User).filter(User.id == int(user_id)).first()
        if not current_user:
            from fastapi import HTTPException
            raise HTTPException(status_code=401, detail="用户不存在")
    else:
        from fastapi import HTTPException
        raise HTTPException(status_code=401, detail="请提供token")

    output = build_rollup_excel(db, dept_code, year, period_type)
    filename = f"rollup_{dept_code}_{year}_{period_type}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/company-report")
def export_company(
    year: int = Query(...),
    report_type: str = Query("monthly"),
    month: int = Query(None),
    period_type: str = Query("actual"),
    token: str = Query(None),
    db: Session = Depends(get_db),
):
    from app.utils.auth import decode_access_token
    from app.models.models import User, Department

    if token:
        payload = decode_access_token(token)
        if not payload:
            from fastapi import HTTPException
            raise HTTPException(status_code=401, detail="Token无效")
        user_id = payload.get("sub")
        current_user = db.query(User).filter(User.id == int(user_id)).first()
        if not current_user:
            from fastapi import HTTPException
            raise HTTPException(status_code=401, detail="用户不存在")

    dept_codes = [d.code for d in db.query(Department).filter(Department.level.in_([2, 3]), Department.code != "company").all()]
    _, subject_list = calc_dept_summary(db, dept_codes, year, month or 1, period_type)

    wb = Workbook()
    ws = wb.active
    ws.title = "商管公司汇总"

    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.cell(row=1, column=1, value="科目代码").font = header_font_white
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).border = thin_border
    ws.cell(row=1, column=2, value="科目名称").font = header_font_white
    ws.cell(row=1, column=2).fill = header_fill
    ws.cell(row=1, column=2).border = thin_border
    ws.cell(row=1, column=3, value="类别").font = header_font_white
    ws.cell(row=1, column=3).fill = header_fill
    ws.cell(row=1, column=3).border = thin_border
    ws.cell(row=1, column=4, value="金额").font = header_font_white
    ws.cell(row=1, column=4).fill = header_fill
    ws.cell(row=1, column=4).border = thin_border

    for i, s in enumerate(subject_list):
        row = i + 2
        ws.cell(row=row, column=1, value=s["code"]).border = thin_border
        ws.cell(row=row, column=2, value=s["name"]).border = thin_border
        ws.cell(row=row, column=3, value=s["category"]).border = thin_border
        cell = ws.cell(row=row, column=4, value=s["value"])
        cell.border = thin_border
        cell.number_format = '#,##0.00'

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"company_summary_{year}_{month or 'all'}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/dept-report")
def export_dept(
    dept_code: str = Query(...),
    year: int = Query(...),
    month: int = Query(...),
    period_type: str = Query("actual"),
    token: str = Query(None),
    db: Session = Depends(get_db),
):
    from app.utils.auth import decode_access_token
    from app.models.models import User, Department

    if token:
        payload = decode_access_token(token)
        if not payload:
            from fastapi import HTTPException
            raise HTTPException(status_code=401, detail="Token无效")

    _, subject_list = calc_dept_summary(db, [dept_code], year, month, period_type)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{dept_code}核算表"

    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.cell(row=1, column=1, value="科目代码").font = header_font_white
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).border = thin_border
    ws.cell(row=1, column=2, value="科目名称").font = header_font_white
    ws.cell(row=1, column=2).fill = header_fill
    ws.cell(row=1, column=2).border = thin_border
    ws.cell(row=1, column=3, value="类别").font = header_font_white
    ws.cell(row=1, column=3).fill = header_fill
    ws.cell(row=1, column=3).border = thin_border
    ws.cell(row=1, column=4, value="金额").font = header_font_white
    ws.cell(row=1, column=4).fill = header_fill
    ws.cell(row=1, column=4).border = thin_border

    for i, s in enumerate(subject_list):
        row = i + 2
        ws.cell(row=row, column=1, value=s["code"]).border = thin_border
        ws.cell(row=row, column=2, value=s["name"]).border = thin_border
        ws.cell(row=row, column=3, value=s["category"]).border = thin_border
        cell = ws.cell(row=row, column=4, value=s["value"])
        cell.border = thin_border
        cell.number_format = '#,##0.00'

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"dept_report_{dept_code}_{year}_{month}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
