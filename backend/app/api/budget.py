from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
from app.database import get_db
from app.models.models import BudgetItem, Department, Subject, Period, FormData, FormItem
from app.utils.auth import require_role, get_current_user

router = APIRouter(prefix="/api/budget", tags=["年度预算"])


class BudgetRowInput(BaseModel):
    subject_code: str
    jan: float = 0
    feb: float = 0
    mar: float = 0
    apr: float = 0
    may: float = 0
    jun: float = 0
    jul: float = 0
    aug: float = 0
    sep: float = 0
    oct: float = 0
    nov: float = 0
    dec: float = 0


class BudgetSaveRequest(BaseModel):
    department_id: int
    year: int
    rows: list[BudgetRowInput]


@router.get("/{dept_id}")
def get_budget(
    dept_id: int,
    year: int = Query(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if current_user.role in ("dept_filler", "dept_reviewer"):
        if current_user.department_id != dept_id:
            raise HTTPException(status_code=403, detail="无权查看其他部门的预算")

    dept = db.query(Department).filter(Department.id == dept_id).first()
    if not dept:
        raise HTTPException(status_code=404, detail="部门不存在")

    subjects = db.query(Subject).filter(Subject.department_id == dept_id, Subject.is_calculated == False).order_by(Subject.sort_order).all()

    budget_map = {}
    items = db.query(BudgetItem).filter(
        BudgetItem.department_id == dept_id,
        BudgetItem.year == year,
    ).all()
    for b in items:
        key = (b.subject_id, b.month)
        budget_map[key] = b.value

    rows = []
    for s in subjects:
        months = {}
        for m in range(1, 13):
            months[m] = budget_map.get((s.id, m), 0)
        rows.append({
            "subject_id": s.id,
            "subject_code": s.code,
            "subject_name": s.name,
            "category": s.category,
            "unit": s.unit,
            "months": months,
        })

    return {"department_name": dept.name, "year": year, "rows": rows}


@router.put("/{dept_id}")
def save_budget(
    dept_id: int,
    req: BudgetSaveRequest,
    db: Session = Depends(get_db),
    current_user=Depends(require_role("super_admin")),
):
    dept = db.query(Department).filter(Department.id == dept_id).first()
    if not dept:
        raise HTTPException(status_code=404, detail="部门不存在")

    db.query(BudgetItem).filter(
        BudgetItem.department_id == dept_id,
        BudgetItem.year == req.year,
    ).delete()

    month_map = {1: "jan", 2: "feb", 3: "mar", 4: "apr", 5: "may", 6: "jun",
                 7: "jul", 8: "aug", 9: "sep", 10: "oct", 11: "nov", 12: "dec"}

    count = 0
    for row in req.rows:
        subject = db.query(Subject).filter(
            Subject.department_id == dept_id,
            Subject.code == row.subject_code,
            Subject.is_calculated == False,
        ).first()
        if not subject:
            continue
        for m in range(1, 13):
            val = getattr(row, month_map[m], 0)
            if val is None:
                val = 0
            db.add(BudgetItem(department_id=dept_id, subject_id=subject.id, year=req.year, month=m, value=float(val)))
            count += 1

    db.commit()
    return {"message": f"保存成功，共写入 {count} 条预算数据"}


@router.post("/sync-to-forms/{year}")
def sync_budget_to_forms(
    year: int,
    db: Session = Depends(get_db),
    current_user=Depends(require_role("super_admin")),
):
    budgets = db.query(BudgetItem).filter(BudgetItem.year == year).all()
    if not budgets:
        raise HTTPException(status_code=400, detail=f"{year}年尚无预算数据，请先录入预算")

    created = 0
    for month in range(1, 13):
        period = db.query(Period).filter(Period.year == year, Period.month == month, Period.period_type == "budget").first()
        if not period:
            period = Period(year=year, month=month, period_type="budget", status="open")
            db.add(period)
            db.flush()

        month_budgets = [b for b in budgets if b.month == month]
        dept_ids = set(b.department_id for b in month_budgets)

        for dept_id in dept_ids:
            dept_budgets = [b for b in month_budgets if b.department_id == dept_id]
            existing = db.query(FormData).filter(
                FormData.period_id == period.id,
                FormData.department_id == dept_id,
            ).first()
            if existing:
                continue

            form = FormData(period_id=period.id, department_id=dept_id, status="draft", filled_by=current_user.id)
            db.add(form)
            db.flush()
            for b in dept_budgets:
                db.add(FormItem(form_id=form.id, subject_id=b.subject_id, value=b.value))
            created += 1

    db.commit()
    return {"message": f"同步完成，共创建 {created} 个预算表单"}


@router.get("/template/{dept_id}")
def download_template(
    dept_id: int,
    year: int = Query(...),
    token: str = Query(None),
    db: Session = Depends(get_db),
):
    from app.utils.auth import decode_access_token
    from app.models.models import User

    if token:
        payload = decode_access_token(token)
        if not payload:
            raise HTTPException(status_code=401, detail="Token无效")
        user_id = payload.get("sub")
        current_user = db.query(User).filter(User.id == int(user_id)).first()
        if not current_user or current_user.role != "super_admin":
            raise HTTPException(status_code=403, detail="仅管理员可下载模板")
    else:
        raise HTTPException(status_code=401, detail="请提供token")
    dept = db.query(Department).filter(Department.id == dept_id).first()
    if not dept:
        raise HTTPException(status_code=404, detail="部门不存在")

    subjects = (
        db.query(Subject)
        .filter(Subject.department_id == dept_id, Subject.is_calculated == False)
        .order_by(Subject.sort_order)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f"{dept.name}"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    headers = ["科目编码", "科目名称", "类别", "单位"]
    for m in range(1, 13):
        headers.append(f"{m}月")
    headers.append("年度合计")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for i, s in enumerate(subjects):
        row = i + 2
        ws.cell(row=row, column=1, value=s.code).border = thin_border
        ws.cell(row=row, column=2, value=s.name).border = thin_border
        ws.cell(row=row, column=3, value=s.category).border = thin_border
        ws.cell(row=row, column=4, value=s.unit).border = thin_border
        for m in range(1, 13):
            cell = ws.cell(row=row, column=4 + m, value="")
            cell.border = thin_border
            cell.alignment = center_align
            cell.number_format = '#,##0.00'
        sum_cell = ws.cell(row=row, column=17, value="")
        sum_cell.border = thin_border
        sum_cell.alignment = center_align

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 8
    for m in range(1, 14):
        ws.column_dimensions[ws.cell(row=1, column=4 + m).column_letter].width = 13

    ws.freeze_panes = "E2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"budget_template_{dept.code}_{year}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/import/{dept_id}")
def import_budget(
    dept_id: int,
    year: int = Query(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(require_role("super_admin")),
):
    dept = db.query(Department).filter(Department.id == dept_id).first()
    if not dept:
        raise HTTPException(status_code=404, detail="部门不存在")

    subjects = db.query(Subject).filter(
        Subject.department_id == dept_id,
        Subject.is_calculated == False,
    ).all()
    subject_map = {s.code: s.id for s in subjects}

    try:
        content = file.file.read()
        wb = load_workbook(BytesIO(content))
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="无法读取Excel文件，请确保格式正确")

    db.query(BudgetItem).filter(
        BudgetItem.department_id == dept_id,
        BudgetItem.year == year,
    ).delete()

    count = 0
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        code = str(row[0]).strip()
        if code not in subject_map:
            skipped += 1
            continue
        subject_id = subject_map[code]
        for m in range(1, 13):
            col_idx = 4 + m
            val = row[col_idx - 1] if col_idx - 1 < len(row) else None
            if val is None or val == "" or val == "":
                val = 0
            try:
                val = float(val)
            except (ValueError, TypeError):
                val = 0
            db.add(BudgetItem(
                department_id=dept_id,
                subject_id=subject_id,
                year=year,
                month=m,
                value=val,
            ))
            count += 1

    if count == 0:
        raise HTTPException(status_code=400, detail="未找到有效数据，请确认科目编码列在A列且与系统一致")

    db.commit()
    return {"message": f"导入成功，共写入 {count} 条记录" + (f"，跳过 {skipped} 个未知科目" if skipped else "")}
